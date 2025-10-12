import docx
import openpyxl
from docxtpl import DocxTemplate
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell

NULLABLE = {"null": True, "blank": True}

# doc = DocxTemplate("../title.docx")
#
#
# context = {"title": ""}
#
# doc.render(context)
# doc.save("123.docx")

# wb = load_workbook("../Excel.xlsx")
# rows = wb.active
# r_list = []
# value_keys = []
# for keys in rows.iter_rows(max_row=1, values_only=True):
#     for k in keys:
#         if k:
#             value_keys.append(str(k))
# for r in rows.iter_rows(min_row=2, values_only=True):
#     r_dict = {}
#     value_v = []
#     for value in r:
#         value_v.append(value)
#         r_dict.update(zip(value_keys, value_v))
#     r_list.append(r_dict)
# print(len(r_list))
# print(r_list)
